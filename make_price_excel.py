import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ราคาสินค้า"

# ข้อมูลทั้งหมด
products = [
    ("🐱 Cat Food",         "Cat Food x10,000",       10,    30),
    ("🐱 Cat Food",         "Cat Food x20,000",       19,    45),
    ("🐱 Cat Food",         "Cat Food x30,000",       25,    65),
    ("🐱 Cat Food",         "Cat Food x45,000",       35,    90),
    ("⚡ XP",               "XP x99,999,999",         25,   480),
    ("🎫 Normal Ticket",    "Normal Ticket x100",     10,    19),
    ("🎫 Normal Ticket",    "Normal Ticket x500",     45,    89),
    ("🎫 Normal Ticket",    "Normal Ticket x1000",    85,   169),
    ("🎫 Normal Ticket",    "Normal Ticket x2999",   239,   469),
    ("🎫 Rare Ticket",      "Rare Ticket x50",        15,    29),
    ("🎫 Rare Ticket",      "Rare Ticket x100",       29,    52),
    ("🎫 Rare Ticket",      "Rare Ticket x200",       50,    99),
    ("🎫 Rare Ticket",      "Rare Ticket x299",       70,   139),
    ("🥊 Battle Item",      "Battle Item x100",       10,    19),
    ("🥊 Battle Item",      "Battle Item x500",       43,    85),
    ("🥊 Battle Item",      "Battle Item x5000",     399,   779),
    ("🥊 Battle Item",      "Battle Item x9,999",    769,  1499),
    ("💰 NP",               "NP x100",                 5,     9),
    ("💰 NP",               "NP x1,000",              49,    99),
    ("💰 NP",               "NP x5,000",             239,   479),
    ("💰 NP",               "NP x9,999",             449,   899),
    ("🔷 Platinum Shard",   "Platinum Shard x10",     30,    59),
    ("🔷 Platinum Shard",   "Platinum Shard x30",     83,   159),
    ("🔷 Platinum Shard",   "Platinum Shard x60",    160,   299),
    ("🔷 Platinum Shard",   "Platinum Shard x99",    250,   489),
    ("📈 Leadership",       "Leadership x100",        20,    39),
    ("📈 Leadership",       "Leadership x1,000",     185,   359),
    ("📈 Leadership",       "Leadership x5,000",     869,  1699),
    ("📈 Leadership",       "Leadership x9,999",    1639,  3199),
    ("💠 Catseye",          "Catseye x100",           20,    39),
    ("💠 Catseye",          "Catseye x500",           87,   169),
    ("💠 Catseye",          "Catseye x5000",         799,  1559),
    ("💠 Catseye",          "Catseye x9,999",       1559,  3039),
    ("🍎 Catfruit",         "Catfruit x100",          30,    50),
    ("🍎 Catfruit",         "Catfruit x300",          85,   130),
    ("🍎 Catfruit",         "Catfruit x500",         139,   210),
    ("🍎 Catfruit",         "Catfruit x998",         269,   399),
    ("🧪 Catamins",         "Catamins x100",          20,    39),
    ("🧪 Catamins",         "Catamins x500",          87,   169),
    ("🧪 Catamins",         "Catamins x5000",        799,  1559),
    ("🧪 Catamins",         "Catamins x9,999",      1559,  3039),
    ("🏆 Legend Ticket",    "Legend Ticket x1",       40,    79),
    ("🏆 Legend Ticket",    "Legend Ticket x2",       74,   145),
    ("🏆 Legend Ticket",    "Legend Ticket x3",      102,   199),
    ("🏆 Legend Ticket",    "Legend Ticket x4",      125,   245),
    ("🎲 Lucky Ticket",     "Lucky Ticket x100",      10,    19),
    ("🎲 Lucky Ticket",     "Lucky Ticket x500",      45,    89),
    ("🎲 Lucky Ticket",     "Lucky Ticket x1000",     85,   169),
    ("🎲 Lucky Ticket",     "Lucky Ticket x2999",    239,   469),
]

# Styles
header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill  = PatternFill("solid", fgColor="1E3A5F")
cat_font     = Font(name="Calibri", bold=True, size=10)
cat_fill     = PatternFill("solid", fgColor="1A1A2E")
cat_font_c   = Font(name="Calibri", bold=True, color="7EB8F7", size=10)
normal_font  = Font(name="Calibri", size=10)
price_font   = Font(name="Calibri", bold=True, color="22C55E", size=10)
old_font     = Font(name="Calibri", color="9CA3AF", size=10, strike=True)
center       = Alignment(horizontal="center", vertical="center")
left         = Alignment(horizontal="left",   vertical="center")
thin         = Side(style="thin", color="D1D5DB")
border       = Border(left=thin, right=thin, top=thin, bottom=thin)

# Header row
ws.row_dimensions[1].height = 24
headers = ["หมวดหมู่", "ชื่อสินค้า", "ราคาขาย (฿)", "ราคาเต็ม (฿)", "ส่วนลด (฿)", "ส่วนลด (%)"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = center
    cell.border    = border

# Data rows
prev_cat = None
for row_idx, (category, name, price, old_price) in enumerate(products, 2):
    ws.row_dimensions[row_idx].height = 20
    discount_amt = old_price - price
    discount_pct = round(discount_amt / old_price * 100)

    is_new_cat = category != prev_cat
    prev_cat   = category

    data = [category, name, price, old_price, discount_amt, f"{discount_pct}%"]
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.border    = border
        cell.alignment = center if col != 2 else left

        if col == 1:
            cell.font = cat_font_c
            cell.fill = PatternFill("solid", fgColor="0F172A")
        elif col == 2:
            cell.font = normal_font
            cell.fill = PatternFill("solid", fgColor="111827" if row_idx % 2 == 0 else "1F2937")
        elif col == 3:
            cell.font = price_font
            cell.fill = PatternFill("solid", fgColor="111827" if row_idx % 2 == 0 else "1F2937")
        elif col == 4:
            cell.font = old_font
            cell.fill = PatternFill("solid", fgColor="111827" if row_idx % 2 == 0 else "1F2937")
        else:
            cell.font = Font(name="Calibri", size=10, color="F59E0B")
            cell.fill = PatternFill("solid", fgColor="111827" if row_idx % 2 == 0 else "1F2937")

# Column widths
col_widths = [22, 28, 16, 16, 14, 12]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze header
ws.freeze_panes = "A2"

out = r"c:\Users\modpo\OneDrive\เดสก์ท็อป\btc\BattleCats\ราคาสินค้า BattleCats.xlsx"
wb.save(out)
print("saved:", out)
